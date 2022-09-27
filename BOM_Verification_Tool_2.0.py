# 실행 전 BOM을 닫아주세요
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
import re

address = input("검증이 필요한 BOM의 저장경로를 입력해주세요(파일명과 확장자명 포함): ")
wb = openpyxl.load_workbook(address)

ws = wb[wb.sheetnames[0]]
ws2 = wb[wb.sheetnames[1]]

# 원본보다 BOM의 행 수가 더 적기 때문에 행 수로 BOM, 원본 시트 판별
if ws.max_row > ws2.max_row:
    ws = wb[wb.sheetnames[1]]
    ws2 = wb[wb.sheetnames[0]]


# list 두 개를 비교해서 겹치는 원소 제외
class Mylist:
    def __init__(self, args):
        self.input = args
        
    def __sub__(self, other):
        return [item for item in self.input if item not in other.input]


# 입력 문자열에서 숫자가 아닌 것들 제거
def str2num(string):
    numbers = re.sub(r'[^0-9]', '', string)
    numbers = int(numbers)
    return numbers


# 알고리즘 적용을 위한 formatting
def formatting(names):
    # 필요 없는 공백 삭제
    names = names.split(" ")
    names = [i for i in names if i != ""]
    names = " ".join(names)

    # hyphen 앞뒤로 blank 있을 시 blank 삭제
    if "- " in names:
        names = names.replace("- ", "-")
    if " -" in names:
        names = names.replace(" -", "-")
    return names


# 자재 수 카운터
def counter(names):
    count = 0
    temp_list = names.split(" ")
    for item in temp_list:
        if "-" in item:
            temp_list2 = item.split("-")
            temp_sum = str2num(temp_list2[1]) - str2num(temp_list2[0]) + 1
            count += temp_sum
        else:
            count += 1
    return count


# hyphen 형식을 blank 형식으로 변환
def hyphen2space(element_list):
    output_list = []
    for h2s in element_list:
        # -을 포함하는 경우 -를 기준으로 split 해서 다시 리스트 결성
        temp_list = []
        if "-" in h2s:
            a = h2s.split("-")
            for temp in a:
                temp_list.append(temp)
            # - 를 blank 형식으로 확장

            alpha = re.findall('[a-zA-Z]',temp_list[0])
            for i in range(str2num(temp_list[1]) - str2num(temp_list[0]) + 1):
                output_element = str2num(temp_list[0]) + i
                output_element = "".join(alpha) + str(output_element)
                output_list.append(output_element)

        # 그렇지 않은 경우
        else:
            output_list.append(h2s)

    return output_list


def fail_comment():
    ws.cell(row=1, column=last_col + 1).value = "수량 불일치 : ("
    ws.cell(row=1, column=last_col + 1).font = Font(size=15, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=2, column=last_col + 1).value = "BOM상 Qty: " + str(BOM_qty)
    ws.cell(row=2, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=3, column=last_col + 1).value = "Schematic상 Qty: " + str(schematic_qty)
    ws.cell(row=3, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=4, column=last_col + 1).value = "누락된 자재: " + " ".join(missing_elements)
    ws.cell(row=4, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')
    ws.cell(row=5, column=last_col + 1).value = "출처 불명확 자재: " + " ".join(unknown_elements)
    ws.cell(row=5, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')


def pass_comment():
    ws.cell(row=1, column=last_col + 1).value = "수량 일치 : )"
    ws.cell(row=1, column=last_col + 1).font = Font(size=15, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=2, column=last_col + 1).value = "BOM상 Qty: " + str(BOM_qty)
    ws.cell(row=2, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=3, column=last_col + 1).value = "Schematic상 Qty: " + str(schematic_qty)
    ws.cell(row=3, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=4, column=last_col + 1).value = ""
    ws.cell(row=5, column=last_col + 1).value = ""


# _________________________________________main_________________________________________ #

# ________________________________________BOM Qty_______________________________________ #

# index 식별 루프
flag = 0  # 이중루프 break 용도
for i in range(1, 5):
    for row in ws[i]:
        index = row.value
        last_col = row.column + 1
        if index == "PASS" or index == "FAIL":
            last_col = row.column
            flag = 1
            break

        if type(index) is str:
            index = index.casefold()
        if index == "Reference".casefold():
            Ref_index = row.column - 1
            index_row = row.row
        if index == "Qty".casefold():
            Qty_index = row.column - 1
    if flag == 1:
        break

# BOM count
BOM_qty = 0
BOM_list = []
for row in ws.rows:
    Names = row[Ref_index].value
    if row[0].row > index_row:
        if Names is not None:
            Names = formatting(Names)
            qty_counted = counter(Names)
            BOM_qty += qty_counted
            BOM_list.extend(Names.split(" "))

            # 결과 출력
            Checking_row = row[Qty_index].row
            if row[Qty_index].value != qty_counted:
                ws.cell(row=Checking_row, column=last_col).value = "FAIL, 입력된 수량: " + str(qty_counted)
                ws.cell(row=Checking_row, column=last_col).font = Font(size=11, name='맑은 고딕', color='00FF0000')
                ws.cell(row=Checking_row, column=Qty_index + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')
                ws.cell(row=Checking_row, column=Ref_index + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')

            else:
                ws.cell(row=Checking_row, column=last_col).value = "PASS"
                ws.cell(row=Checking_row, column=last_col).font = Font(size=11, name='맑은 고딕', color='000000FF')
                ws.cell(row=Checking_row, column=Qty_index + 1).font = Font(size=11, name='맑은 고딕', color='000000FF')
                ws.cell(row=Checking_row, column=Ref_index + 1).font = Font(size=11, name='맑은 고딕', color='000000FF')

# 자재 중복 여부 검사
overlap_dict = {}
overlap_elements = []
for temp in BOM_list:
    if temp in overlap_dict:
        overlap_dict[temp] += 1
    else:
        overlap_dict[temp] = 1

    if overlap_dict[temp] == 2:
        overlap_elements.append(temp)

# ______________________________________Schematic Qty______________________________________ #

# index 식별 루프
for i in range(1, 5):
    for row in ws2[i]:
        temp = row.value
        if type(temp) is str:
            temp = temp.casefold()
        if temp == "VALUE".casefold():
            Val_index_ws2 = row.column - 1
            index_row_ws2 = row.row
        if temp == "Qty".casefold():
            Qty_index_ws2 = row.column - 1
        if temp == "Reference".casefold():
            Ref_index_ws2 = row.column - 1

# Schematic count
schematic_qty = 0
schematic_list = []
for row in ws2.rows:
    if row[0].row > index_row_ws2:
        if str(row[Val_index_ws2].value) not in ("BM", "DNI", "TP", "MTH", "IND_FDCL", "PAD") and "581Pin" not in str(
                row[Val_index_ws2].value):
            schematic_qty += row[Qty_index_ws2].value  # Schematic상 Qty 값 종합
            Ref_value_ws2 = formatting(row[Ref_index_ws2].value)  # 양식에 맞춰 fomatting
            schematic_list.extend(Ref_value_ws2.split(" "))  # 각각의 물품번호를 하나의 리스트에 통합

# ______________________________________최종 결과________________________________________ #

missing_elements = Mylist(schematic_list) - Mylist(BOM_list)
unknown_elements = Mylist(BOM_list) - Mylist(schematic_list)
final_missing = Mylist(hyphen2space(missing_elements)) - Mylist(hyphen2space(unknown_elements))
final_unknown = Mylist(hyphen2space(unknown_elements)) - Mylist(hyphen2space(missing_elements))

print(missing_elements)
print(unknown_elements)
print(hyphen2space(missing_elements))
print(hyphen2space(unknown_elements))
print(final_missing)
print(final_unknown)

# ______________________________________결과 출력________________________________________ #

print("BOM상 Qty:", BOM_qty)
print("Schematic상 Qty:", schematic_qty)
# print("Schematic_list: ", " ".join(schematic_list))
# print("BOM_list: ", " ".join(BOM_list))

# 에러 발생
if BOM_qty != schematic_qty or missing_elements or unknown_elements or overlap_elements:
    if final_missing or final_unknown:
        print("누락된 자재가 있습니다:", " ".join(final_missing))
        print("출처 불명확 자재가 있습니다:", " ".join(final_unknown))
        print("중복 입력된 자재가 있습니다:", " ".join(overlap_elements))
        fail_comment()
    else:
        pass_comment()

else:
    pass_comment()

wb.save(address)


