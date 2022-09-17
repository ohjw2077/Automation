import os
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\user\Downloads\220907_MT2737_이형 소자LIST - 복사본.xlsx')
ws = wb["전체소자"]
ws2 = wb['데이터시트 수급현황']
print(ws, ws2)

for i in ws.columns:
    for j in range(5):
        if i[j].value == "Datasheet":
            datasheet_val, datasheet_start = i, j + 1
        if i[j].value == "품명":
            name_val, name_start = i, j + 1


original_datasheet = os.listdir(r'C:\Users\user\Downloads\MTDatasheet')

datasheet_list =[]
for i in range(len(original_datasheet)):
    name = original_datasheet[i]
    ws2.cell(row=i+1, column=1).value = name
    name = name.upper()[:-4]
    datasheet_list.append(name)



for i in range(name_start, ws.max_row):
    flag = 0
    if name_val[i].value != None:
        name2 = name_val[i].value
        name2 = name2.upper()
        for j in datasheet_list:
            if name2 in j:
                print("found:", name2)
                ws.cell(row=datasheet_val[i].row, column=datasheet_val[i].column).value = "O"
                flag = 1
                break
        if flag == 1:
            continue
        print("not found:", name2)
        ws.cell(row=datasheet_val[i].row, column=datasheet_val[i].column).value = "X"

wb.save(r'C:\Users\user\Downloads\220907_MT2737_이형 소자LIST - 복사본.xlsx')
