import openpyxl

wb1 = openpyxl.load_workbook(r"D:\Project\5G\SA515M\(Restricted)E850_ITT_38EA_RF_20220310_copied.xlsx")
ws1 = wb1["LAM-E870"]

wb2 = openpyxl.load_workbook(r"D:\Project\5G\SA515M\LTE.xlsx", data_only=True)
ws2 = wb2["Max Power"]


#Detecting a Dual SIM Format
Sim_Slot_Detector = ws1["F3"].value
if Sim_Slot_Detector[0:3] == "SIM2":
    SIM_MODE = 1
else:
    SIM_MODE = 0


#####################################Import data from LTE.xlsx#####################################

#Detect the number of devices
num_of_devices = 0
temp = 0
for row in ws2.rows:
    if type(row[0].value) == int:
        temp = 1
        num_of_devices += 1
    elif type(row[0].value) == str and temp == 1:
        break
#print(num_of_devices)

#Detect and Export MAX POWER data only
temp2 = 0
MAX_POWER_list = []
for row in ws2.rows:
    if row[0].value == "Result :":
        checkpoint = row[0].row + 1
        channel_value = ws2.cell(row = checkpoint, column = 2).value
        if channel_value[6:8] == "12":  # 12RB
            for j in range(3):
                MAX_POWER_list2 = []
                channel_value = ws2.cell(row=checkpoint, column=2 + (j * 2)).value
                MAX_POWER_list2.append(channel_value[:5]) #append channel value as the first element in each list
                #print("MAX POWER, Channel", channel_value[:6])
                for k in range(num_of_devices):
                    MAX_POWER = ws2.cell(row = checkpoint + 3 + k, column = 2 + (j * 2)).value
                    #print("시료 %d: %s" %(k+1,MAX_POWER))
                    MAX_POWER_list2.append(MAX_POWER)

                MAX_POWER_list.append(MAX_POWER_list2)

# for a in range(len(MAX_POWER_list)):
#      print(MAX_POWER_list[a][0])

###################Attach data to (Restricted)E850_ITT_38EA_RF_20220310_copied.xlsx####################

#initial setup
for row2 in ws1.rows:
    if row2[0].value == "Channel":
        ws1_channel_row = row2
    elif row2[1].value == "#1":
        ws1_starting_row = row2[1].row

#Import MAX POWER data
y=0
for x in range(len(ws1_channel_row)):
    ws1_channel = str(ws1_channel_row[x].value) #채널 행 정보 str로 변환 후 할당
    if ws1_channel[0:4] == "SIM1" and ws1_channel[7:] == MAX_POWER_list[y][0]:

        if y < len(MAX_POWER_list) - 1:
            y += 1

        for z in range(1, len(MAX_POWER_list[y])):
            MAX_POWER_IMPORT = MAX_POWER_list[y][z]

            # z - 1은 1번부터 시작해서, x + 1은 python loop 시작은 0인데 excel row 시작은 1부터라서


wb1.save(r"D:\Project\5G\SA515M\(Restricted)E850_ITT_38EA_RF_20220310_copied.xlsx")