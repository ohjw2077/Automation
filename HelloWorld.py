import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\AutoPython\Qty_check_E871L.xlsx")
ws = wb[wb.sheetnames[0]]
ws2 = wb[wb.sheetnames[1]]
print(ws)
print(ws2)

print(wb.sheetnames)



# wb.save(r"C:\Users\user\Documents\AutoPython\Qty_check_E871L.xlsx")