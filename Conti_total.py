import openpyxl
import ast

wb_init = openpyxl.load_workbook(r"D:\Conti\Conti_BOM\Total_BOM\total_list.xlsx", data_only=True)
ws_init = wb_init["Sheet1"]
wb_main = openpyxl.load_workbook(r"D:\Conti\Conti_BOM\Total_BOM\BOM_total.xlsx")
ws_main = wb_main["Sheet1"]

name_list = []

for i in range(40):
    designated_col = i + 7
    filename = ws_init.cell(row=i + 2, column=2).value
    filename2 = filename + "_released_assy"
    filename3 = filename2 + ".xlsx"
    file_address = "D:\\Conti\\Conti_BOM\\Total_BOM\\" + filename3

    print(f"______________________{filename}____________________")
    wb = openpyxl.load_workbook(file_address, data_only=True)
    ws = wb[filename2]

    ws_main.cell(row=1, column=designated_col).value = filename

    b = 0
    for row in ws.rows:
        if row[0].value == "레벨":
            b += 1
            if b == 4:
                starting_row = row[0].row + 1
                break

    for row2 in ws.rows:
        if row2[0].row >= starting_row:
            qty = ast.literal_eval(row2[8].value)
            name = row2[2].value
            print(name)
            if name not in name_list:
                name_list.append(name)
                x = len(name_list) + 1
                ws_main.cell(row=x, column=1).value = name
                ws_main.cell(row=x, column=2).value = row2[5].value
                ws_main.cell(row=x, column=3).value = row2[14].value
                ws_main.cell(row=x, column=4).value = row2[15].value
                ws_main.cell(row=x, column=5).value = row2[21].value
                ws_main.cell(row=x, column=6).value = row2[22].value
                ws_main.cell(row=x, column=designated_col).value = qty
            else:
                y = name_list.index(name) + 2
                ws_main.cell(row=y, column=designated_col).value = qty


wb_main.save(r"D:\Conti\Conti_BOM\Total_BOM\BOM_total.xlsx")







