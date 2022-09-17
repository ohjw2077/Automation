import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\AUTO_BOM.xlsx")
ws = wb["Sheet1"]

#열 정보 불러오기
for col in ws.columns:
    for i in range(3):
        if col[i].value == "VALUE":
            value_col = chr(col[i].column + 64)
        if col[i].value == "Qty":
            qty_col = chr(col[i].column + 64)


# BM, DNI 등등 필요없는 값 삭제
for VALUE in ws[value_col]:
    if str(VALUE.value) in ("BM", "DNI", "TP", "MTH", "IND_FDCL") or "581Pin" in str(VALUE.value):
        delete_me = VALUE.row
        ws.delete_rows(delete_me,1)

# wb.save(r"C:\Users\user\Documents\AUTOBOM.xlsx")

#필요 없는 값 제외 총 자재 개수
total_qty = 0
for qty in ws[qty_col]:
    if type(qty.value) is int:
        total_qty += qty.value

print(total_qty)

#중복 값 합치기




