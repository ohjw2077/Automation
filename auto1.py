import openpyxl as op
class Mylist(list):
    def __init__(self, *args):
        super(Mylist, self).__init__(args)

    def __sub__(self, other):
        return self.__class__(*[x for x in self if x not in other])

wb = op.load_workbook(r"C:\Users\user\Documents\openpyxl_test.xlsx")

ws_deleted = Mylist()
ws_list = Mylist()
ws_comp = wb.sheetnames
for i in range(len(ws_comp)):
    ws_list.append(ws_comp[i])

for empty_sheets in ws_list:
    ws = wb["%s" % empty_sheets]
    if ws["A1"].value is None:
        ws_deleted.append(empty_sheets)
        wb.remove(ws)

ws_left = ws_list - ws_deleted
print("Following sheets are deleted", ws_deleted)
print("remaining sheets:",ws_left )

wb.save(r"C:\Users\user\Documents\openpyxl_test.xlsx")




