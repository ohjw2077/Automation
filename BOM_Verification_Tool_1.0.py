# 실행 전 BOM을 닫아주세요
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
import re

address = input("검증이 필요한 BOM의 저장경로를 입력해주세요(파일명과 확장자명 포함): ")
wb = openpyxl.load_workbook(address)
ws = wb[wb.sheetnames[0]]

# 입력 문자열에서 숫자가 아닌 것들 제거
def str2num(string):
    numbers = re.sub(r'[^0-9]', '', string)
    numbers = int(numbers)
    return numbers


# 알고리즘 적용을 위한 formatting
def formatting(names):
    # 필요 없는 공백 삭제
    names = names.split(" ")
    names = [i for i in names if i != ""]
    names = " ".join(names)

    # hyphen 앞뒤로 blank 있을 시 blank 삭제
    if "- " in names:
        names = names.replace("- ", "-")
    if " -" in names:
        names = names.replace(" -", "-")
    return names


# hyphen 형식을 blank 형식으로 변환
def hyphen2space(element_list):
    output_list = []
    for h2s in element_list:
        # -을 포함하는 경우 -를 기준으로 split 해서 다시 리스트 결성
        temp_list = []
        if "-" in h2s:
            a = h2s.split("-")
            for temp in a:
                temp_list.append(temp)

            # 물품번호에서 문자만 추출
            alpha = re.findall('[a-zA-Z]', temp_list[0])
            for i in range(str2num(temp_list[1]) - str2num(temp_list[0]) + 1):
                output_element = str2num(temp_list[0]) + i
                output_element = "".join(alpha) + str(output_element)
                output_list.append(output_element)

        # 그렇지 않은 경우
        else:
            output_list.append(h2s)

    return output_list


def fail_comment():
    ws.cell(row=1, column=last_col + 1).value = "수량 불일치 : ("
    ws.cell(row=1, column=last_col + 1).font = Font(size=15, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=2, column=last_col + 1).value = "BOM상 Qty: " + str(BOM_qty)
    ws.cell(row=2, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=3, column=last_col + 1).value = "Schematic상 Qty: " + str(schematic_qty)
    ws.cell(row=3, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000', bold=True)
    ws.cell(row=4, column=last_col + 1).value = "누락된 자재: " + " ".join(missing_elements)
    ws.cell(row=4, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')
    ws.cell(row=5, column=last_col + 1).value = "출처 불명확 자재: " + " ".join(unknown_elements)
    ws.cell(row=5, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')


def pass_comment():
    ws.cell(row=1, column=last_col + 1).value = "수량 일치 : )"
    ws.cell(row=1, column=last_col + 1).font = Font(size=15, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=2, column=last_col + 1).value = "BOM상 Qty: " + str(BOM_qty)
    ws.cell(row=2, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=3, column=last_col + 1).value = "Schematic상 Qty: " + str(schematic_qty)
    ws.cell(row=3, column=last_col + 1).font = Font(size=11, name='맑은 고딕', color='000000FF', bold=True)
    ws.cell(row=4, column=last_col + 1).value = ""
    ws.cell(row=5, column=last_col + 1).value = ""


# _________________________________________main_________________________________________ #

# ________________________________________BOM Qty_______________________________________ #

# 각 항목 인덱스 식별
flag = 0  # 이중루프 break 용도
for i in range(1, 5):
    for row in ws[i]:
        index = row.value

        if type(index) is str:
            index = index.casefold()  # 대소문자 관계 없이 식별

        if index == "Reference".casefold():
            Ref_index = row.column - 1
            index_row = row.row

        if index == "Qty".casefold():
            Qty_index = row.column - 1

    if flag == 1:
        break

# 물품개수 카운팅 시작
BOM_qty = 0
BOM_list = []
for row in ws.rows:
    Names = row[Ref_index].value
    if row[0].row > index_row:
        if Names is not None:
            Names = formatting(Names)
            Names = Names.split(" ")
            Names = hyphen2space(Names)
            qty_counted = len(Names)
            BOM_qty += qty_counted
            BOM_list.extend(Names)

            # BOM상의 Qty값과 알고리즘으로 산출된 갯수를 비교하여 결과 출력
            Checking_row = row[Qty_index].row
            if row[Qty_index].value != qty_counted:
                ws.cell(row=Checking_row, column=Qty_index + 2).value = "FAIL, 현재 입력된 수량은 " + str(qty_counted) + "개 입니다."
                ws.cell(row=Checking_row, column=Qty_index + 2).font = Font(size=11, name='맑은 고딕', color='00FF0000')
                ws.cell(row=Checking_row, column=Qty_index + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')
                ws.cell(row=Checking_row, column=Ref_index + 1).font = Font(size=11, name='맑은 고딕', color='00FF0000')

            else:
                ws.cell(row=Checking_row, column=Qty_index + 2).value = "PASS"
                ws.cell(row=Checking_row, column=Qty_index + 2).font = Font(size=11, name='맑은 고딕', color='000000FF')
                ws.cell(row=Checking_row, column=Qty_index + 1).font = Font(size=11, name='맑은 고딕', color='000000FF')
                ws.cell(row=Checking_row, column=Ref_index + 1).font = Font(size=11, name='맑은 고딕', color='000000FF')

# 자재 중복 여부 검사
overlap_dict = {}
overlap_elements = []
for temp in BOM_list:
    if temp in overlap_dict:
        overlap_dict[temp] += 1
    else:
        overlap_dict[temp] = 1

    if overlap_dict[temp] == 2:
        overlap_elements.append(temp)

ws.cell(row=1, column=Qty_index + 2).value = "중복 입력된 자재가 있습니다: " + " ".join(overlap_elements)
ws.cell(row=1, column=Qty_index + 2).font = Font(size=11, name='맑은 고딕', color='00FF0000')

wb.save(address)